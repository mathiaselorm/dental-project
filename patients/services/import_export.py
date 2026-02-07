# patients/services/import_export.py

"""
Patient Import/Export Services.

Provides robust CSV/Excel import and export functionality with:
- Validation and error handling
- Duplicate detection
- Support for updating existing records (by folder_number)
- Progress tracking with PatientImportExportJob
- Audit logging
- Template generation (CSV/Excel)
- Error report generation (CSV/Excel)

Notes:
- Patient folder_number is immutable and system-generated. Imports NEVER create with a supplied folder_number.
- Defaults for patient_type and acquisition_channel are handled in services.patients.register_patient().
  (Typically: Regular/General and WALK_IN, seeded & active.)
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any, BinaryIO, Callable, Generator, Optional
from uuid import UUID

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from accounts.audit import audit_log

from ..models import Gender, Patient, PatientImportExportJob, normalize_ghana_phone
from .patients import (
    PatientValidationError,
    find_potential_duplicates,
    register_patient,
    register_patient_with_duplicate_check,
    update_patient,
)

logger = logging.getLogger(__name__)


# =========================
# Column mappings
# =========================

# Raw header -> internal field name
_IMPORT_COLUMN_MAPPING_RAW = {
    # Primary identifiers
    "folder_number": "folder_number",
    "folder number": "folder_number",
    "patient_id": "folder_number",
    "patient id": "folder_number",

    # Names
    "first_name": "first_name",
    "first name": "first_name",
    "firstname": "first_name",
    "given_name": "first_name",
    "given name": "first_name",

    "last_name": "last_name",
    "last name": "last_name",
    "lastname": "last_name",
    "surname": "last_name",
    "family_name": "last_name",
    "family name": "last_name",

    # Demographics
    "date_of_birth": "date_of_birth",
    "date of birth": "date_of_birth",
    "dob": "date_of_birth",
    "birth_date": "date_of_birth",
    "birth date": "date_of_birth",
    "birthdate": "date_of_birth",

    "gender": "gender",
    "sex": "gender",

    # Contact
    "phone_number": "phone_number",
    "phone number": "phone_number",
    "phone": "phone_number",
    "mobile": "phone_number",
    "mobile_number": "phone_number",
    "mobile number": "phone_number",
    "telephone": "phone_number",
    "tel": "phone_number",

    "email": "email",
    "email_address": "email",
    "email address": "email",
    "e-mail": "email",

    # Address & Details
    "address": "address",
    "home_address": "address",
    "home address": "address",

    "occupation": "occupation",
    "job": "occupation",
    "profession": "occupation",

    "nationality": "nationality",
    "country": "nationality",

    # Patient Type (by name)
    "patient_type": "patient_type_name",
    "patient type": "patient_type_name",
    "type": "patient_type_name",
    "category": "patient_type_name",

    # Acquisition Channel (by code OR name-like input; we normalize later)
    "acquisition_channel": "acquisition_channel_code",
    "acquisition channel": "acquisition_channel_code",
    "channel": "acquisition_channel_code",
    "source": "acquisition_channel_code",
    "how_did_you_hear": "acquisition_channel_code",
    "how did you hear": "acquisition_channel_code",

    # Child & Guardian
    "is_child": "is_child",
    "is child": "is_child",
    "child": "is_child",
    "minor": "is_child",

    "guardian_name": "guardian_name",
    "guardian name": "guardian_name",
    "guardian": "guardian_name",
    "parent_name": "guardian_name",
    "parent name": "guardian_name",

    "guardian_phone": "guardian_phone",
    "guardian phone": "guardian_phone",
    "guardian_mobile": "guardian_phone",
    "guardian mobile": "guardian_phone",
    "parent_phone": "guardian_phone",
    "parent phone": "guardian_phone",

    "guardian_email": "guardian_email",
    "guardian email": "guardian_email",
    "parent_email": "guardian_email",
    "parent email": "guardian_email",

    "guardian_occupation": "guardian_occupation",
    "guardian occupation": "guardian_occupation",

    "guardian_address": "guardian_address",
    "guardian address": "guardian_address",

    # Referral
    "referrer_name": "referrer_name",
    "referrer name": "referrer_name",
    "referred_by": "referrer_name",
    "referred by": "referrer_name",

    # Optional: support referrer_patient_id if a clinic exports/uses it later
    "referrer_patient_id": "referrer_patient_id",
    "referrer patient id": "referrer_patient_id",

    # Profile
    "profile_picture_url": "profile_picture_url",
    "profile_picture": "profile_picture_url",
    "profile picture": "profile_picture_url",
    "avatar_url": "profile_picture_url",
    "avatar": "profile_picture_url",
}


def _normalize_header(header: Any) -> str:
    """
    Normalize header text so we can map things like:
      "First Name*" -> "first name"
      "Date of Birth* (YYYY-MM-DD)" -> "date of birth"
    """
    h = "" if header is None else str(header)
    h = h.strip()

    # remove BOM, asterisks, and parenthetical hints
    h = h.replace("\ufeff", "")
    h = h.replace("*", "")
    h = re.sub(r"\(.*?\)", "", h)

    # normalize separators/spaces
    h = h.replace("-", " ").replace("_", " ")
    h = re.sub(r"\s+", " ", h).strip().lower()

    return h


# Build normalized mapping once
IMPORT_COLUMN_MAPPING: dict[str, str] = {
    _normalize_header(k): v for k, v in _IMPORT_COLUMN_MAPPING_RAW.items()
}


# Export columns in order (internal field -> label)
EXPORT_COLUMNS = [
    ("folder_number", "Folder Number"),
    ("first_name", "First Name"),
    ("last_name", "Last Name"),
    ("date_of_birth", "Date of Birth"),
    ("age", "Age"),
    ("gender", "Gender"),
    ("phone_number", "Phone Number"),
    ("email", "Email"),
    ("address", "Address"),
    ("occupation", "Occupation"),
    ("nationality", "Nationality"),
    ("patient_type", "Patient Type"),  # name
    ("acquisition_channel", "Acquisition Channel"),  # CODE (for re-import)
    ("is_child", "Is Child"),
    ("guardian_name", "Guardian Name"),
    ("guardian_phone", "Guardian Phone"),
    ("guardian_email", "Guardian Email"),
    ("guardian_occupation", "Guardian Occupation"),
    ("guardian_address", "Guardian Address"),
    ("profile_picture_url", "Profile Picture URL"),
    ("created_at", "Created At"),
    ("updated_at", "Updated At"),
]

# Required fields for import (patient_type + acquisition_channel can default via services)
REQUIRED_IMPORT_FIELDS = {"first_name", "last_name", "date_of_birth", "gender", "phone_number"}


# =========================
# Data Classes
# =========================

@dataclass
class ImportOptions:
    """Configuration options for patient import."""
    skip_duplicates: bool = False         # Skip rows that match existing patients
    update_existing: bool = False         # Update existing patients by folder_number
    duplicate_threshold: int = 50         # Score threshold for duplicate detection
    skip_invalid_rows: bool = True        # Continue on validation errors
    dry_run: bool = False                # Validate only, don't persist
    batch_size: int = 100                # Commit transaction every N rows
    generate_error_report: bool = True   # Create an error report file when job is provided


@dataclass
class ExportOptions:
    """Configuration options for patient export."""
    include_deleted: bool = False
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    patient_type_ids: list[UUID] = field(default_factory=list)
    acquisition_channel_ids: list[UUID] = field(default_factory=list)
    search_query: Optional[str] = None


@dataclass
class ImportRowResult:
    """Result of processing a single import row."""
    row_number: int
    success: bool
    action: str = ""  # created|updated|skipped|validated|error
    patient_id: Optional[UUID] = None
    folder_number: Optional[str] = None
    errors: dict = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportResult:
    """Overall result of an import operation."""
    total_rows: int = 0
    processed_rows: int = 0
    created_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    row_results: list[ImportRowResult] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# =========================
# Parsing Utilities
# =========================

def _map_headers(headers: list[Any]) -> dict[int, str]:
    """
    Map column indices to internal field names.
    Returns: {col_index -> field_name}
    """
    mapping: dict[int, str] = {}
    for idx, header in enumerate(headers):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        if normalized in IMPORT_COLUMN_MAPPING:
            mapping[idx] = IMPORT_COLUMN_MAPPING[normalized]
            continue

        # Fallback: if header equals export internal field names
        export_fields = {f for f, _ in EXPORT_COLUMNS}
        if normalized in export_fields:
            mapping[idx] = normalized

    return mapping


def _parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    # Excel serial dates (numbers)
    if isinstance(value, (int, float)):
        try:
            excel_epoch = date(1899, 12, 30)
            return excel_epoch + timedelta(days=int(value))
        except Exception:
            return None

    value_str = str(value).strip()
    if not value_str:
        return None

    date_formats = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%d.%m.%Y",
    ]
    for fmt in date_formats:
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue
    return None


def _parse_boolean(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value

    s = str(value).strip().lower()
    if s in ("true", "yes", "y", "1", "t", "on"):
        return True
    if s in ("false", "no", "n", "0", "f", "off"):
        return False
    return None


def _parse_gender(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None

    s = str(value).strip().lower()
    if s in ("male", "m", "man", "boy"):
        return Gender.MALE
    if s in ("female", "f", "woman", "girl"):
        return Gender.FEMALE

    for choice_value, _label in Gender.choices:
        if s == str(choice_value).lower():
            return choice_value

    return None


def _normalize_phone_number(value: Any) -> Optional[str]:
    """
    Normalize to Ghana +233XXXXXXXXX format.
    Delegates to the model's normalize function for consistency.
    """
    if value is None or value == "":
        return None
    return normalize_ghana_phone(str(value).strip())


def _clean_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _coerce_uuid(value: Any) -> Optional[UUID]:
    if value is None or value == "":
        return None
    try:
        return UUID(str(value))
    except Exception:
        return None


def _parse_import_row(row_data: dict[str, Any], row_number: int) -> tuple[dict[str, Any], list[str]]:
    """
    Parse and validate a single import row.
    Returns: (parsed_data, warnings)
    Raises: PatientValidationError(errors_dict)
    """
    parsed: dict[str, Any] = {}
    warnings: list[str] = []
    errors: dict[str, Any] = {}

    for field, value in row_data.items():
        if field == "date_of_birth":
            d = _parse_date(value)
            if value and not d:
                errors[field] = f"Invalid date format: {value}"
            else:
                parsed[field] = d

        elif field == "gender":
            g = _parse_gender(value)
            if value and not g:
                errors[field] = f"Invalid gender: {value}. Use 'Male' or 'Female'."
            else:
                parsed[field] = g

        elif field == "phone_number":
            p = _normalize_phone_number(value)
            if value and not p:
                errors[field] = f"Invalid phone number: {value}"
            else:
                parsed[field] = p

        elif field == "guardian_phone":
            parsed[field] = _normalize_phone_number(value) if value else None

        elif field == "is_child":
            b = _parse_boolean(value)
            parsed[field] = b if b is not None else False

        elif field in ("email", "guardian_email"):
            email = _clean_string(value)
            if email:
                email = email.lower()
                if "@" not in email or "." not in email:
                    errors[field] = f"Invalid email format: {email}"
                else:
                    parsed[field] = email
            else:
                parsed[field] = None

        elif field in ("patient_type_name", "acquisition_channel_code", "referrer_name"):
            parsed[field] = _clean_string(value)

        elif field == "referrer_patient_id":
            u = _coerce_uuid(value)
            if value and not u:
                errors[field] = f"Invalid UUID: {value}"
            else:
                parsed[field] = u

        else:
            parsed[field] = _clean_string(value)

    # Required fields
    for required in REQUIRED_IMPORT_FIELDS:
        if not parsed.get(required):
            errors[required] = "This field is required."

    # Child rules
    if parsed.get("is_child"):
        if not parsed.get("guardian_name"):
            errors["guardian_name"] = "Guardian name is required for child patients."
        if not parsed.get("guardian_phone"):
            errors["guardian_phone"] = "Guardian phone is required for child patients."

    if errors:
        raise PatientValidationError(errors)

    return parsed, warnings


def _find_existing_patient_by_folder(folder_number: str) -> Optional[Patient]:
    if not folder_number:
        return None
    return Patient.all_objects.filter(folder_number=folder_number).first()


# =========================
# Readers
# =========================

def read_csv_file(file_obj: BinaryIO, encoding: str = "utf-8-sig") -> Generator[tuple[list[str], list[list[Any]]], None, None]:
    """
    Read CSV file -> yields (headers, rows_as_lists).
    (We read into memory to know total rows; typical clinic imports are manageable.)
    """
    file_obj.seek(0)
    raw = file_obj.read()

    # decode with fallback
    text_content = None
    for enc in (encoding, "utf-8", "latin-1", "cp1252", "iso-8859-1"):
        try:
            text_content = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if text_content is None:
        raise ValueError("Could not decode CSV. Please upload a UTF-8 encoded file.")

    if text_content.startswith("\ufeff"):
        text_content = text_content[1:]

    reader = csv.reader(io.StringIO(text_content))
    rows = list(reader)
    if not rows:
        yield [], []
        return

    yield rows[0], rows[1:]


def read_excel_file(file_obj: BinaryIO, sheet_name: Optional[str] = None) -> Generator[tuple[list[str], list[list[Any]]], None, None]:
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl is required for Excel imports. Install it with: pip install openpyxl") from e

    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        yield [], []
        return

    headers = [str(h) if h is not None else "" for h in rows[0]]

    data_rows: list[list[Any]] = []
    for row in rows[1:]:
        if all(cell is None or cell == "" for cell in row):
            continue
        data_rows.append(list(row))

    yield headers, data_rows


def detect_file_format(filename: str) -> str:
    ext = os.path.splitext(filename.lower())[1]
    if ext in (".xlsx", ".xls", ".xlsm"):
        return PatientImportExportJob.FileFormat.EXCEL
    if ext in (".csv", ".txt"):
        return PatientImportExportJob.FileFormat.CSV
    raise ValueError(f"Unsupported file format: {ext}")


# =========================
# Import
# =========================

def import_patients_from_file(
    file_obj: BinaryIO,
    filename: str,
    options: ImportOptions,
    job: Optional[PatientImportExportJob] = None,
    request=None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> ImportResult:
    result = ImportResult()

    file_format = detect_file_format(filename)
    reader = read_csv_file(file_obj) if file_format == PatientImportExportJob.FileFormat.CSV else read_excel_file(file_obj)

    try:
        headers, data_rows = next(reader)
    except StopIteration:
        headers, data_rows = [], []

    if not headers:
        result.errors.append({"detail": "File is empty or has no headers."})
        return result

    column_mapping = _map_headers(headers)
    if not column_mapping:
        result.errors.append({
            "detail": (
                "No recognizable columns found. Expected columns like: "
                "first name, last name, date of birth, gender, phone number."
            )
        })
        return result

    mapped_fields = set(column_mapping.values())
    missing_required = REQUIRED_IMPORT_FIELDS - mapped_fields
    if missing_required:
        result.warnings.append(
            f"Missing columns for required fields: {', '.join(sorted(missing_required))}. "
            "Rows missing these values will fail validation."
        )

    result.total_rows = len(data_rows)

    if job:
        job.total_rows = result.total_rows
        job.save(update_fields=["total_rows", "updated_at"])

    for row_idx, row_values in enumerate(data_rows):
        row_number = row_idx + 2  # account for header row
        row_result = ImportRowResult(row_number=row_number, success=False)

        try:
            # Map row values -> internal fields
            row_data: dict[str, Any] = {}
            for col_idx, field_name in column_mapping.items():
                if col_idx < len(row_values):
                    row_data[field_name] = row_values[col_idx]

            if not any(v not in (None, "") for v in row_data.values()):
                row_result.action = "skipped"
                row_result.success = True
                row_result.warnings.append("Empty row")
                result.skipped_count += 1
                continue

            # Parse row
            parsed_data, parse_warnings = _parse_import_row(row_data, row_number)
            row_result.warnings.extend(parse_warnings)

            # Extract folder_number for update matching; NEVER pass into create payload
            folder_number = parsed_data.pop("folder_number", None)

            # Find existing for update
            existing_patient = None
            if folder_number and options.update_existing:
                existing_patient = _find_existing_patient_by_folder(folder_number)
                if not existing_patient:
                    row_result.warnings.append(f"folder_number '{folder_number}' not found; row will create a new patient.")

            # Referral extraction (do this before duplicate check so payload is complete)
            referrer_name = parsed_data.pop("referrer_name", None)
            referrer_patient_id = parsed_data.pop("referrer_patient_id", None)

            if referrer_name or referrer_patient_id:
                referral_payload: dict[str, Any] = {}
                if referrer_name:
                    referral_payload["referrer_name"] = referrer_name
                if referrer_patient_id:
                    referral_payload["referrer_patient_id"] = referrer_patient_id
                parsed_data["referral"] = referral_payload

            if options.dry_run:
                # For dry run, still do non-atomic duplicate check for validation
                if not existing_patient:
                    duplicates = find_potential_duplicates(
                        first_name=parsed_data.get("first_name"),
                        last_name=parsed_data.get("last_name"),
                        phone_number=parsed_data.get("phone_number"),
                        date_of_birth=parsed_data.get("date_of_birth"),
                        email=parsed_data.get("email"),
                        threshold=options.duplicate_threshold,
                    )
                    if duplicates:
                        top = duplicates[0]
                        msg = (
                            f"Potential duplicate found: folder_number={top.patient.folder_number}, "
                            f"match_score={top.score}"
                        )
                        if options.skip_duplicates:
                            row_result.action = "skipped"
                            row_result.success = True
                            row_result.warnings.append(f"Would skip - {msg}")
                            result.skipped_count += 1
                            continue
                        row_result.warnings.append(msg)
                row_result.success = True
                row_result.action = "validated"
                continue

            # Persist row with atomic duplicate checking
            if existing_patient:
                with transaction.atomic():
                    updated = update_patient(patient=existing_patient, payload=parsed_data, request=request)
                    row_result.patient_id = updated.id
                    row_result.folder_number = updated.folder_number
                    row_result.action = "updated"
                    row_result.success = True
                    result.updated_count += 1
            else:
                # Use atomic duplicate check + registration to prevent race conditions
                dup_result = register_patient_with_duplicate_check(
                    payload=parsed_data,
                    duplicate_threshold=options.duplicate_threshold,
                    skip_if_duplicate=options.skip_duplicates,
                    request=request,
                )
                
                if dup_result.has_duplicates:
                    # Duplicates found and skip_duplicates=True
                    top = dup_result.duplicates[0]
                    msg = (
                        f"Potential duplicate found: folder_number={top.patient.folder_number}, "
                        f"match_score={top.score}"
                    )
                    row_result.action = "skipped"
                    row_result.success = True
                    row_result.warnings.append(f"Skipped - {msg}")
                    result.skipped_count += 1
                else:
                    # Patient created successfully
                    created = dup_result.created_patient
                    row_result.patient_id = created.id
                    row_result.folder_number = created.folder_number
                    row_result.action = "created"
                    row_result.success = True
                    result.created_count += 1

        except PatientValidationError as e:
            row_result.errors = e.errors
            row_result.action = "error"
            result.error_count += 1
            if not options.skip_invalid_rows:
                result.errors.append({"row": row_number, "errors": e.errors})
                break

        except Exception as e:
            logger.exception("Error processing import row %s", row_number)
            row_result.errors = {"detail": "An unexpected error occurred while processing this row."}
            row_result.action = "error"
            result.error_count += 1
            if not options.skip_invalid_rows:
                result.errors.append({"row": row_number, "errors": {"detail": "An unexpected error occurred."}})
                break

        finally:
            result.processed_rows += 1
            result.row_results.append(row_result)

            if progress_callback:
                progress_callback(result.processed_rows, result.total_rows)

            if job and result.processed_rows % 10 == 0:
                job.processed_rows = result.processed_rows
                job.success_count = result.created_count + result.updated_count
                job.error_count = result.error_count
                job.skip_count = result.skipped_count
                job.save(update_fields=["processed_rows", "success_count", "error_count", "skip_count", "updated_at"])

    # Final job update
    if job:
        job.processed_rows = result.processed_rows
        job.success_count = result.created_count + result.updated_count
        job.error_count = result.error_count
        job.skip_count = result.skipped_count

        # Store per-row errors for UI
        job.row_errors = [
            {"row": rr.row_number, "errors": rr.errors, "warnings": rr.warnings, "action": rr.action}
            for rr in result.row_results
            if rr.errors
        ]
        job.save(update_fields=["processed_rows", "success_count", "error_count", "skip_count", "row_errors", "updated_at"])

    # Audit log (only if not dry_run)
    if not options.dry_run:
        audit_log(
            action="PATIENT_BULK_IMPORT",
            request=request,
            success=(result.error_count == 0),
            metadata={
                "total_rows": result.total_rows,
                "created": result.created_count,
                "updated": result.updated_count,
                "skipped": result.skipped_count,
                "errors": result.error_count,
                "filename": filename,
                "update_existing": options.update_existing,
                "skip_duplicates": options.skip_duplicates,
            },
        )

    return result


# =========================
# Job wrappers (optional but production-friendly)
# =========================

def _job_storage_path(prefix: str, filename: str) -> str:
    """
    Returns a safe path relative to storage (MEDIA_ROOT for default storage).
    """
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", filename).strip("_")
    return f"patients/{prefix}/{safe}"


def run_import_job(
    *,
    job: PatientImportExportJob,
    file_obj: BinaryIO,
    filename: str,
    options: ImportOptions,
    request=None,
) -> ImportResult:
    """
    Convenience wrapper that marks job started/completed/failed and (optionally)
    stores an error report file in job.result_file.
    """
    try:
        job.mark_started()
        result = import_patients_from_file(
            file_obj=file_obj,
            filename=filename,
            options=options,
            job=job,
            request=request,
        )

        # Save error report if needed
        if options.generate_error_report and result.error_count > 0:
            error_ext = ".xlsx" if detect_file_format(filename) == PatientImportExportJob.FileFormat.EXCEL else ".csv"
            report_name = _job_storage_path("import_errors", f"{job.id}_errors{error_ext}")

            buf = io.BytesIO()
            generate_import_error_report(result=result, output_file=buf, file_format=detect_file_format(filename))
            buf.seek(0)

            saved_path = default_storage.save(report_name, ContentFile(buf.read()))
            job.result_file = saved_path
            job.save(update_fields=["result_file", "updated_at"])

        job.mark_completed()
        return result

    except Exception as e:
        logger.exception("Import job %s failed", job.id)
        # Sanitize error message before storing
        error_msg = "An unexpected error occurred during import."
        job.mark_failed(error_msg)
        raise


# =========================
# Export
# =========================

def _patient_to_export_row(patient: Patient) -> dict[str, Any]:
    """
    Export Acquisition Channel as CODE (so re-import is easy),
    Patient Type as NAME.
    """
    return {
        "folder_number": patient.folder_number,
        "first_name": patient.first_name,
        "last_name": patient.last_name,
        "date_of_birth": patient.date_of_birth.isoformat() if patient.date_of_birth else "",
        "age": patient.age,
        "gender": patient.gender,
        "phone_number": patient.phone_number,
        "email": patient.email or "",
        "address": patient.address or "",
        "occupation": patient.occupation or "",
        "nationality": patient.nationality or "",
        "patient_type": patient.patient_type.name if patient.patient_type else "",
        "acquisition_channel": patient.acquisition_channel.code if patient.acquisition_channel else "",
        "is_child": "Yes" if patient.is_child else "No",
        "guardian_name": patient.guardian_name or "",
        "guardian_phone": patient.guardian_phone or "",
        "guardian_email": patient.guardian_email or "",
        "guardian_occupation": patient.guardian_occupation or "",
        "guardian_address": patient.guardian_address or "",
        "profile_picture_url": patient.profile_picture_url or "",
        "created_at": patient.created_at.isoformat() if patient.created_at else "",
        "updated_at": patient.updated_at.isoformat() if patient.updated_at else "",
    }


def _build_export_queryset(options: ExportOptions):
    qs = Patient.all_objects.all() if options.include_deleted else Patient.objects.all()
    qs = qs.select_related("patient_type", "acquisition_channel")

    if options.date_from:
        qs = qs.filter(created_at__date__gte=options.date_from)
    if options.date_to:
        qs = qs.filter(created_at__date__lte=options.date_to)

    if options.patient_type_ids:
        qs = qs.filter(patient_type_id__in=options.patient_type_ids)
    if options.acquisition_channel_ids:
        qs = qs.filter(acquisition_channel_id__in=options.acquisition_channel_ids)

    if options.search_query:
        q = options.search_query.strip()
        qs = qs.filter(
            Q(folder_number__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(phone_number__icontains=q)
            | Q(email__icontains=q)
        )

    return qs.order_by("folder_number")


def export_patients_to_csv(
    output_file: BinaryIO,
    options: ExportOptions,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> int:
    qs = _build_export_queryset(options)
    total_count = qs.count()

    # BOM for Excel compatibility
    output_file.write(b"\xef\xbb\xbf")

    wrapper = io.TextIOWrapper(output_file, encoding="utf-8", newline="")
    writer = csv.writer(wrapper)

    writer.writerow([label for _field, label in EXPORT_COLUMNS])

    for idx, patient in enumerate(qs.iterator(chunk_size=500), start=1):
        row_data = _patient_to_export_row(patient)
        row = [row_data.get(field, "") for field, _label in EXPORT_COLUMNS]
        writer.writerow(row)

        if progress_callback and idx % 100 == 0:
            progress_callback(idx, total_count)

    wrapper.flush()
    wrapper.detach()
    return total_count


def export_patients_to_excel(
    output_file: BinaryIO,
    options: ExportOptions,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> int:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl") from e

    qs = _build_export_queryset(options)
    total_count = qs.count()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    headers = [label for _field, label in EXPORT_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_idx = 2
    for idx, patient in enumerate(qs.iterator(chunk_size=500), start=1):
        row_data = _patient_to_export_row(patient)
        for col_idx, (field, _label) in enumerate(EXPORT_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
        row_idx += 1

        if progress_callback and idx % 100 == 0:
            progress_callback(idx, total_count)

    # Auto-width
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    wb.save(output_file)
    return total_count


def export_patients(
    output_file: BinaryIO,
    file_format: str,
    options: ExportOptions,
    job: Optional[PatientImportExportJob] = None,
    request=None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> int:
    def track_progress(processed: int, total: int):
        if progress_callback:
            progress_callback(processed, total)
        if job and processed % 100 == 0:
            job.processed_rows = processed
            job.total_rows = total
            job.save(update_fields=["processed_rows", "total_rows", "updated_at"])

    if file_format == PatientImportExportJob.FileFormat.CSV:
        count = export_patients_to_csv(output_file, options, track_progress)
    elif file_format == PatientImportExportJob.FileFormat.EXCEL:
        count = export_patients_to_excel(output_file, options, track_progress)
    else:
        raise ValueError(f"Unsupported export format: {file_format}")

    audit_log(
        action="PATIENT_BULK_EXPORT",
        request=request,
        success=True,
        metadata={
            "count": count,
            "format": file_format,
            "include_deleted": options.include_deleted,
            "date_from": str(options.date_from) if options.date_from else None,
            "date_to": str(options.date_to) if options.date_to else None,
        },
    )
    return count


def run_export_job(
    *,
    job: PatientImportExportJob,
    output_format: str,
    options: ExportOptions,
    request=None,
) -> str:
    """
    Creates an export file in storage and saves path to job.result_file.
    Returns saved file path.
    """
    try:
        job.mark_started()

        ext = ".xlsx" if output_format == PatientImportExportJob.FileFormat.EXCEL else ".csv"
        out_name = _job_storage_path("exports", f"{job.id}_patients{ext}")

        buf = io.BytesIO()
        count = export_patients(
            output_file=buf,
            file_format=output_format,
            options=options,
            job=job,
            request=request,
        )
        buf.seek(0)

        saved_path = default_storage.save(out_name, ContentFile(buf.read()))
        job.result_file = saved_path
        job.total_rows = count
        job.processed_rows = count
        job.mark_completed()
        job.save(update_fields=["result_file", "total_rows", "processed_rows", "updated_at"])
        return saved_path

    except Exception as e:
        logger.exception("Export job %s failed", job.id)
        # Sanitize error message before storing
        error_msg = "An unexpected error occurred during export."
        job.mark_failed(error_msg)
        raise


# =========================
# Template Generation
# =========================

def generate_import_template_csv(output_file: BinaryIO) -> None:
    output_file.write(b"\xef\xbb\xbf")

    wrapper = io.TextIOWrapper(output_file, encoding="utf-8", newline="")
    writer = csv.writer(wrapper)

    headers = [
        "First Name*",
        "Last Name*",
        "Date of Birth* (YYYY-MM-DD)",
        "Gender* (Male/Female)",
        "Phone Number* (+233XXXXXXXXX)",
        "Email",
        "Address",
        "Occupation",
        "Nationality",
        "Patient Type (optional: defaults to Regular/General)",
        "Acquisition Channel (optional: defaults to WALK_IN)",
        "Is Child (Yes/No)",
        "Guardian Name",
        "Guardian Phone",
        "Guardian Email",
        "Guardian Occupation",
        "Guardian Address",
        "Referred By (name)",
    ]
    writer.writerow(headers)

    example = [
        "John",
        "Doe",
        "1990-05-15",
        "Male",
        "+233241234567",
        "john.doe@example.com",
        "123 Main Street, Accra",
        "Engineer",
        "Ghanaian",
        "Regular",
        "WALK_IN",
        "No",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    writer.writerow(example)

    wrapper.flush()
    wrapper.detach()


def generate_import_template_excel(output_file: BinaryIO) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as e:
        raise ImportError("openpyxl is required for Excel template generation.") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Import"

    headers = [
        ("First Name", True),
        ("Last Name", True),
        ("Date of Birth", True),
        ("Gender", True),
        ("Phone Number", True),
        ("Email", False),
        ("Address", False),
        ("Occupation", False),
        ("Nationality", False),
        ("Patient Type", False),
        ("Acquisition Channel", False),
        ("Is Child", False),
        ("Guardian Name", False),
        ("Guardian Phone", False),
        ("Guardian Email", False),
        ("Guardian Occupation", False),
        ("Guardian Address", False),
        ("Referred By", False),
    ]

    required_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    optional_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    header_font = Font(bold=True)

    for col_idx, (header, required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=f"{header}*" if required else header)
        cell.font = header_font
        cell.fill = required_fill if required else optional_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Validations
    gender_dv = DataValidation(type="list", formula1='"Male,Female"', allow_blank=False)
    ws.add_data_validation(gender_dv)
    gender_dv.add("D2:D1000")

    child_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(child_dv)
    child_dv.add("L2:L1000")

    ws_help = wb.create_sheet("Instructions")
    instructions = [
        ("Patient Import Instructions", ""),
        ("", ""),
        ("Required Fields", "Fields marked with * are required."),
        ("Defaults", "If Patient Type or Acquisition Channel are blank, defaults apply (Regular/General, WALK_IN)."),
        ("Phone Number", "Ghana format: +233XXXXXXXXX (e.g., +233241234567)."),
        ("Gender", "Male or Female."),
        ("Is Child", "If Yes, guardian name + phone are required."),
        ("Referral", "Only needed if Acquisition Channel is REFERRED."),
    ]
    for row_idx, (c1, c2) in enumerate(instructions, 1):
        ws_help.cell(row=row_idx, column=1, value=c1)
        ws_help.cell(row=row_idx, column=2, value=c2)
        if row_idx == 1:
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True, size=14)

    ws_help.column_dimensions["A"].width = 28
    ws_help.column_dimensions["B"].width = 70

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = "A2"
    wb.save(output_file)


# =========================
# Error Report Generation
# =========================

def generate_import_error_report(result: ImportResult, output_file: BinaryIO, file_format: str) -> None:
    error_rows = [r for r in result.row_results if r.errors]

    if file_format == PatientImportExportJob.FileFormat.CSV:
        output_file.write(b"\xef\xbb\xbf")
        wrapper = io.TextIOWrapper(output_file, encoding="utf-8", newline="")
        writer = csv.writer(wrapper)

        writer.writerow(["Row Number", "Action", "Errors", "Warnings"])
        for rr in error_rows:
            writer.writerow([
                rr.row_number,
                rr.action,
                "; ".join(f"{k}: {v}" for k, v in rr.errors.items()),
                "; ".join(rr.warnings),
            ])

        wrapper.flush()
        wrapper.detach()
        return

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise ImportError("openpyxl required for Excel error reports.") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Errors"

    ws.cell(row=1, column=1, value="Import Error Report").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Total Rows: {result.total_rows}")
    ws.cell(row=3, column=1, value=f"Created: {result.created_count}")
    ws.cell(row=4, column=1, value=f"Updated: {result.updated_count}")
    ws.cell(row=5, column=1, value=f"Skipped: {result.skipped_count}")
    ws.cell(row=6, column=1, value=f"Errors: {result.error_count}")

    headers = ["Row Number", "Action", "Field", "Error Message"]
    header_fill = PatternFill(start_color="FFCDD2", fill_type="solid")

    base_row = 8
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=base_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    row_idx = base_row + 1
    for rr in error_rows:
        for field, err in rr.errors.items():
            ws.cell(row=row_idx, column=1, value=rr.row_number)
            ws.cell(row=row_idx, column=2, value=rr.action)
            ws.cell(row=row_idx, column=3, value=field)
            ws.cell(row=row_idx, column=4, value=str(err))
            row_idx += 1

    wb.save(output_file)
